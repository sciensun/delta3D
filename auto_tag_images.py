import argparse
import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import yaml
from PIL import Image
from tqdm import tqdm


GROUP_COLUMNS = {
    "animal_labels": ("animal_pred", "animal_conf"),
    "material_labels": ("material_pred", "material_conf"),
    "global_style_labels": ("global_style_pred", "global_style_conf"),
}


PROMPT_HINTS = {
    "lion": ["a photo of a lion sculpture", "a stylized carved wooden lion"],
    "deer": ["a photo of a deer sculpture", "a deer figurine with antlers"],
    "dog": ["a photo of a dog sculpture", "a dog figurine"],
    "cat": ["a photo of a cat sculpture", "a cat figurine"],
    "bird": ["a photo of a bird sculpture", "a porcelain bird figurine"],
    "bear": ["a photo of a bear sculpture", "a bear figurine"],
    "horse": ["a photo of a horse sculpture", "a folk art horse sculpture"],
    "ox": ["a photo of an ox sculpture", "a bull or ox figurine"],
    "goat_sheep": ["a photo of a goat or sheep sculpture"],
    "unknown_animal": ["a photo of an animal sculpture"],
    "wood": ["a photo of a wooden animal sculpture", "a carved wood animal figurine"],
    "stone": ["a stone animal sculpture"],
    "porcelain": ["a photo of a porcelain animal figurine"],
    "bronze": ["a bronze animal sculpture"],
    "concrete": ["a concrete animal sculpture"],
    "painted": ["a painted animal sculpture"],
    "scanned_real_object": ["a realistic 3D scan of an animal object"],
    "unknown_material": ["an animal sculpture of unknown material"],
    "realistic_scan": ["a realistic 3D scan of an animal"],
    "classical_sculpture": ["a classical animal sculpture"],
    "stylized_sculpture": ["a stylized animal sculpture"],
    "folk_art": ["a folk art animal sculpture"],
    "toy_like": ["a toy-like animal figurine"],
    "ornament": ["an ornamental animal figurine"],
    "low_poly_or_blocky": ["a blocky faceted animal sculpture"],
    "smooth_porcelain_like": ["a smooth porcelain animal figurine"],
    "groove_carved": ["a sculpture with deep carved grooves", "a carved animal with groove details"],
    "faceted_blocky": ["a blocky faceted animal sculpture"],
    "smooth_sculptural": ["a smooth sculptural animal figurine"],
    "ornamental_relief": ["an ornamental relief animal sculpture"],
    "simplified_anatomy": ["an animal sculpture with simplified anatomy"],
    "exaggerated_proportion": ["an animal sculpture with exaggerated proportions"],
    "base_or_support": ["an animal sculpture standing on a base or support"],
    "realistic_reference": ["a realistic reference animal model"],
}


KEYWORDS = {
    "lion": ["lion", "lowe", "löwe"],
    "deer": ["deer", "hirsch", "stag"],
    "dog": ["dog", "hound"],
    "cat": ["cat"],
    "bird": ["bird", "owl", "eagle", "pigeon"],
    "bear": ["bear"],
    "horse": ["horse"],
    "ox": ["ox", "bull", "cow"],
    "goat_sheep": ["goat", "sheep", "ram", "urial"],
    "wood": ["wood", "wooden", "carved"],
    "stone": ["stone", "statue", "classical"],
    "porcelain": ["porcelain", "ceramic"],
    "bronze": ["bronze", "metal"],
    "concrete": ["concrete"],
    "painted": ["painted", "folk"],
    "scanned_real_object": ["scan", "3d_scan", "artec"],
    "realistic_scan": ["scan", "realistic", "artec"],
    "classical_sculpture": ["classical", "statue", "sculpture"],
    "stylized_sculpture": ["stylized"],
    "folk_art": ["folk"],
    "toy_like": ["toy"],
    "ornament": ["ornament"],
    "low_poly_or_blocky": ["low_poly", "lowpoly", "blocky"],
    "smooth_porcelain_like": ["porcelain", "smooth"],
    "groove_carved": ["carved", "wood"],
    "faceted_blocky": ["low_poly", "lowpoly", "blocky"],
    "smooth_sculptural": ["smooth", "porcelain"],
    "ornamental_relief": ["ornament", "relief"],
    "simplified_anatomy": ["stylized", "folk", "toy"],
    "exaggerated_proportion": ["stylized", "toy"],
    "base_or_support": ["statue", "sculpture"],
    "realistic_reference": ["scan", "realistic"],
}


def ensure_can_write(path, overwrite):
    if path.exists() and not overwrite:
        raise FileExistsError(f"{path} exists. Pass --overwrite to replace it.")
    path.parent.mkdir(parents=True, exist_ok=True)


def load_vocab(path):
    with open(path, "r", encoding="utf-8") as handle:
        return yaml.safe_load(handle)


def make_prompts(label):
    return PROMPT_HINTS.get(label, [f"a photo of a {label.replace('_', ' ')} animal sculpture"])


def try_load_open_clip(device):
    try:
        import torch
        import open_clip
    except Exception as exc:
        print(f"WARNING: open_clip unavailable, using filename fallback: {exc}", file=sys.stderr)
        return None

    try:
        model, _, preprocess = open_clip.create_model_and_transforms("ViT-B-32", pretrained="laion2b_s34b_b79k")
        tokenizer = open_clip.get_tokenizer("ViT-B-32")
        model.to(device).eval()
        return {"torch": torch, "model": model, "preprocess": preprocess, "tokenizer": tokenizer, "device": device}
    except Exception as exc:
        print(f"WARNING: failed to initialize open_clip, using filename fallback: {exc}", file=sys.stderr)
        return None


def clip_scores(clip_state, image_path, labels):
    torch = clip_state["torch"]
    model = clip_state["model"]
    preprocess = clip_state["preprocess"]
    tokenizer = clip_state["tokenizer"]
    device = clip_state["device"]

    prompts = []
    owners = []
    for label in labels:
        for prompt in make_prompts(label):
            prompts.append(prompt)
            owners.append(label)

    try:
        image = preprocess(Image.open(image_path).convert("RGB")).unsqueeze(0).to(device)
        text = tokenizer(prompts).to(device)
        with torch.no_grad():
            image_features = model.encode_image(image)
            text_features = model.encode_text(text)
            image_features = image_features / image_features.norm(dim=-1, keepdim=True)
            text_features = text_features / text_features.norm(dim=-1, keepdim=True)
            prompt_probs = (100.0 * image_features @ text_features.T).softmax(dim=-1).cpu().numpy()[0]
    except Exception as exc:
        print(f"WARNING: CLIP failed on {image_path}, using fallback: {exc}", file=sys.stderr)
        return None

    scores = {label: 0.0 for label in labels}
    for owner, score in zip(owners, prompt_probs):
        scores[owner] = max(scores[owner], float(score))
    total = sum(scores.values()) or 1.0
    return {label: score / total for label, score in scores.items()}


def fallback_scores(text, labels):
    haystack = text.lower().replace("-", "_").replace(" ", "_")
    raw = {}
    for label in labels:
        score = 0.02
        for keyword in KEYWORDS.get(label, []):
            if keyword.lower() in haystack:
                score += 0.35
        if label.startswith("unknown_"):
            score += 0.05
        raw[label] = score
    total = sum(raw.values()) or 1.0
    return {label: score / total for label, score in raw.items()}


def top_label(scores):
    label, score = max(scores.items(), key=lambda item: item[1])
    return label, round(float(score), 4)


def top_tags(scores, count=3):
    ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)[:count]
    return [label for label, _ in ranked], [round(float(score), 4) for _, score in ranked]


def suggest_role(global_style, deformation_tags):
    if global_style == "realistic_scan" or "realistic_reference" in deformation_tags:
        return "target_reference"
    if global_style in {"stylized_sculpture", "folk_art", "ornament", "low_poly_or_blocky", "smooth_porcelain_like"}:
        return "source_style"
    return "mixed"


def score_group(clip_state, image_path, labels, fallback_text):
    if clip_state:
        scores = clip_scores(clip_state, image_path, labels)
        if scores is not None:
            return scores
    return fallback_scores(fallback_text, labels)


def add_thumbnails(xlsx_path, rows):
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
    except Exception as exc:
        print(f"WARNING: cannot insert thumbnails without openpyxl image support: {exc}", file=sys.stderr)
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="thumbnail")
    ws.column_dimensions["A"].width = 18

    for idx, row in enumerate(rows, start=2):
        image_path = row.get("preview_path", "")
        if not image_path or not Path(image_path).exists():
            continue
        try:
            img = XLImage(image_path)
            img.width = 96
            img.height = 96
            ws.row_dimensions[idx].height = 76
            ws.add_image(img, f"A{idx}")
        except Exception as exc:
            print(f"WARNING: failed to add thumbnail for {image_path}: {exc}", file=sys.stderr)

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Generate candidate visual labels for 3D animal previews.")
    parser.add_argument("--manifest", default="outputs/manifest.csv")
    parser.add_argument("--vocab", default="outputs/deformation_vocab.yaml")
    parser.add_argument("--out", default="outputs/labels_auto.xlsx")
    parser.add_argument("--manual_out", default="outputs/labels_for_manual.xlsx")
    parser.add_argument("--jsonl", default="")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--no_thumbnails", action="store_true")
    args = parser.parse_args()

    out_path = Path(args.out)
    manual_path = Path(args.manual_out)
    ensure_can_write(out_path, args.overwrite)
    ensure_can_write(manual_path, args.overwrite)
    if args.jsonl:
        ensure_can_write(Path(args.jsonl), args.overwrite)

    vocab = load_vocab(args.vocab)
    manifest = pd.read_csv(args.manifest).fillna("")

    clip_state = None
    try:
        import torch

        device = "cuda" if torch.cuda.is_available() else "cpu"
    except Exception:
        device = "cpu"
    clip_state = try_load_open_clip(device)

    rows = []
    jsonl_records = []
    for _, item in tqdm(manifest.iterrows(), total=len(manifest), desc="tagging"):
        image_path = item.get("preview_path", "")
        fallback_text = " ".join(str(item.get(col, "")) for col in ["model_id", "model_name", "preview_path", "source_url", "notes"])
        if not Path(image_path).exists():
            print(f"WARNING: preview image missing: {image_path}", file=sys.stderr)

        animal_scores = score_group(clip_state, image_path, vocab["animal_labels"], fallback_text)
        material_scores = score_group(clip_state, image_path, vocab["material_labels"], fallback_text)
        style_scores = score_group(clip_state, image_path, vocab["global_style_labels"], fallback_text)
        deform_scores = score_group(clip_state, image_path, vocab["deformation_tags"], fallback_text)

        animal, animal_conf = top_label(animal_scores)
        material, material_conf = top_label(material_scores)
        style, style_conf = top_label(style_scores)
        deform_tags, deform_conf = top_tags(deform_scores)
        role = suggest_role(style, deform_tags)

        row = {
            "model_id": item.get("model_id", ""),
            "model_name": item.get("model_name", ""),
            "preview_path": image_path,
            "source_url": item.get("source_url", ""),
            "animal_pred": animal,
            "animal_conf": animal_conf,
            "material_pred": material,
            "material_conf": material_conf,
            "global_style_pred": style,
            "global_style_conf": style_conf,
            "deformation_tags_pred": ";".join(deform_tags),
            "deformation_tags_conf": ";".join(str(value) for value in deform_conf),
            "target_role_pred": role,
            "manual_animal": "",
            "manual_material": "",
            "manual_global_style": "",
            "manual_deformation_tags": "",
            "manual_local_parts": "",
            "manual_target_role": "",
            "manual_keep": "",
            "manual_note": "",
        }
        rows.append(row)
        jsonl_records.append(
            {
                **row,
                "scores": {
                    "animal": animal_scores,
                    "material": material_scores,
                    "global_style": style_scores,
                    "deformation_tags": deform_scores,
                },
            }
        )

    labels = pd.DataFrame(rows)
    labels.to_excel(out_path, index=False)
    labels.to_excel(manual_path, index=False)
    if not args.no_thumbnails:
        add_thumbnails(out_path, rows)
        add_thumbnails(manual_path, rows)

    if args.jsonl:
        with open(args.jsonl, "w", encoding="utf-8") as handle:
            for record in jsonl_records:
                handle.write(json.dumps(record, ensure_ascii=False) + "\n")

    print(f"Wrote {len(labels)} rows to {out_path}")
    print(f"Wrote manual review workbook to {manual_path}")
    if args.jsonl:
        print(f"Wrote JSONL scores to {args.jsonl}")


if __name__ == "__main__":
    main()
